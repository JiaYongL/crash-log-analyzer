"""
server.py — Flask API backend for the crash log analyser

Endpoints
---------
POST /analyze
    Accepts multipart file uploads (one or more log files).
    Groups files by their top-level directory path, analyses each group with
    the Ollama-backed CrashLogAnalyzer, and returns a ZIP archive containing:

      crash_analysis_<timestamp>.xlsx   — Excel report (sorted, styled)
      grouped/
        <log_type>/                     — omitted if null/empty
          <exception_type>/             — omitted if null/empty
            <frame_fingerprint>/        — omitted if null/empty
              <original_group_dir>/
                <uuid_subfolder>/
                  <all uploaded files>

GET /health
    Simple liveness check.

Run with:
    python server.py
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import tempfile
import traceback
import zipfile
from collections import defaultdict
from datetime import datetime

from flask import Flask, Response, jsonify, request, send_file

sys.path.insert(0, os.path.dirname(__file__))

from analyzer import CrashLogAnalyzer
from config import DEFAULT_MODEL, LOG_HEAD_LINES
from log_scanner import preferred_files

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 512 * 1024 * 1024  # 512 MB


# ---------------------------------------------------------------------------
# CORS
# ---------------------------------------------------------------------------

@app.after_request
def add_cors_headers(response: Response) -> Response:
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.route("/analyze", methods=["OPTIONS"])
def preflight():
    return "", 204


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.route("/health")
def health():
    return jsonify({"status": "ok", "model": DEFAULT_MODEL})


# ---------------------------------------------------------------------------
# Main analysis endpoint
# ---------------------------------------------------------------------------

@app.route("/analyze", methods=["POST"])
def analyze():
    uploaded = request.files.getlist("files")
    model = request.form.get("model", DEFAULT_MODEL)

    if not uploaded:
        return jsonify({"error": "No files received"}), 400

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_filename = f"crash_analysis_{timestamp}.xlsx"
    zip_filename  = f"crash_analysis_{timestamp}.zip"

    with tempfile.TemporaryDirectory() as tmp_root:
        groups: dict[str, list[str]] = defaultdict(list)

        for file_obj in uploaded:
            # webkitRelativePath is forwarded as the filename with slashes
            rel_path = file_obj.filename.replace("\\", "/")
            dest = os.path.join(tmp_root, rel_path)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            file_obj.save(dest)

            # Group by top-level folder component
            parts = rel_path.split("/")
            top = parts[0] if len(parts) > 1 else "root"
            groups[top].append(dest)

        # Run analysis
        analyzer = CrashLogAnalyzer(model=model)
        rows = _analyze_groups(analyzer, groups, tmp_root)

        # Build Excel
        xlsx_bytes = _build_excel(rows)

        # Build ZIP (Excel + reorganised directory tree) while tmp_root exists
        zip_bytes = _build_grouped_zip(rows, tmp_root, xlsx_bytes, xlsx_filename)

    resp = send_file(
        io.BytesIO(zip_bytes),
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_filename,
    )
    resp.headers["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp

# ---------------------------------------------------------------------------
# Helpers — analysis
# ---------------------------------------------------------------------------

def _analyze_groups(
    analyzer: CrashLogAnalyzer,
    groups: dict[str, list[str]],
    tmp_root: str,
) -> list[dict]:
    """Analyse each file group and return a list of result rows."""
    rows: list[dict] = []

    for group_key, files in groups.items():
        selected = preferred_files(files)
        combined = _merge_heads(selected)

        try:
            result = analyzer.analyze_as_dict(
                combined,
                stream=False,  # server-side: no stdout streaming
                json_mode=True,
            )
        except json.JSONDecodeError:
            result = {"error": "Model did not return valid JSON"}
        except Exception as exc:  # noqa: BLE001
            result = {"error": str(exc)}
            traceback.print_exc()

        rows.append({
            **result,
            "group": group_key,
            "files_analysed": len(selected),
        })

    return rows


def _merge_heads(files: list[str]) -> str:
    """Concatenate the first LOG_HEAD_LINES lines of each file."""
    parts: list[str] = []
    for path in files:
        with open(path, encoding="utf-8", errors="ignore") as fh:
            head = "".join(fh.readlines()[:LOG_HEAD_LINES])
        parts.append(f"\n\n=== File: {os.path.basename(path)} ===\n{head}")
    return "".join(parts)


# ---------------------------------------------------------------------------
# Helpers — sorting
# ---------------------------------------------------------------------------

def _sort_rows(rows: list[dict]) -> list[dict]:
    """
    Sort analysis rows for easier reading in Excel.

    Primary   → log_type       (alphabetical; blank rows last)
    Secondary → exception_type (alphabetical within the same log type)
    Tertiary  → group          (alphabetical directory name)

    Rows that failed analysis (have an "error" key) sink to the bottom.
    """
    def _key(row: dict) -> tuple[str, str, str, str, str]:
        log_type = (row.get("log_type") or "").lower()
        exception_type = (row.get("exception_type") or "").lower()
        frame_type = (row.get("frame_type") or "").lower()
        frame_fingerprint = (row.get("frame_fingerprint") or "").lower()
        current_thread = (row.get("current_thread") or "").lower()
        return (log_type, exception_type, frame_type, frame_fingerprint, current_thread)

    return sorted(rows, key=_key)


# ---------------------------------------------------------------------------
# Helpers — grouped ZIP
# ---------------------------------------------------------------------------

# Characters that are unsafe in cross-platform directory names.
_UNSAFE_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _sanitize_segment(value: str) -> str:
    """
    Strip characters that are unsafe in directory names on Windows/macOS/Linux,
    and truncate to 120 chars so paths don't exceed OS limits.
    """
    safe = _UNSAFE_RE.sub("_", value).strip(". ")
    return safe[:120] if safe else "_"


def _build_grouped_zip(
    rows: list[dict],
    tmp_root: str,
    xlsx_bytes: bytes,
    xlsx_filename: str,
) -> bytes:
    """
    Build an in-memory ZIP that contains:

    1. The Excel report at the archive root.
    2. A ``grouped/`` tree where each original crash directory is nested under
       its analysis result segments::

         grouped/
           <log_type>/              ← omitted when null / empty
             <exception_type>/      ← omitted when null / empty
               <frame_fingerprint>/ ← omitted when null / empty
                 <original_group>/
                   <uuid_subfolder>/
                     file.log
                     ...

    Any segment whose analysis value is null, empty, or whitespace-only is
    skipped so the structure stays clean.
    """
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # ── Excel at archive root ──────────────────────────────────────────
        zf.writestr(xlsx_filename, xlsx_bytes)

        # ── Reorganised directory tree ─────────────────────────────────────
        for row in rows:
            group = (row.get("group") or "").strip()
            if not group:
                continue

            src_dir = os.path.join(tmp_root, group)
            if not os.path.isdir(src_dir):
                # Single-file group or missing directory — skip silently.
                continue

            # Build the prefix path segments, omitting any empty/null values.
            segments = ["grouped", row.get("log_type", "").strip()]
            dir_name = []
            for key in ("exception_type", "frame_fingerprint"):
                raw = (row.get(key) or "").strip()
                if raw:
                    dir_name.append(_sanitize_segment(raw))
            segments.append("+".join(dir_name))  # always include the original group dir name

            prefix = "/".join(segments)  # e.g. "grouped/hs_err/EXCEPTION.../chrome_elf.dll/5.1.1.840_..."

            # Walk the source directory and add every file to the archive.
            for dirpath, _dirs, filenames in os.walk(src_dir):
                for fname in filenames:
                    abs_path = os.path.join(dirpath, fname)

                    # Relative path from tmp_root: "group/uuid/file.log"
                    rel = os.path.relpath(abs_path, tmp_root).replace(os.sep, "/")

                    # Strip the leading group component (already in `segments`)
                    # so we don't double-up the group dir name.
                    rel_parts = rel.split("/")           # ["group", "uuid", "file.log"]
                    suffix    = "/".join(rel_parts[1:])  # "uuid/file.log"

                    arc_path = f"{prefix}/{suffix}"
                    zf.write(abs_path, arc_path)

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers — Excel
# ---------------------------------------------------------------------------

def _build_excel(rows: list[dict]) -> bytes:
    """Build an Excel workbook from the analysis rows and return raw bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Sort rows so the sheet is immediately readable without manual filtering.
    rows = _sort_rows(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crash Analysis"

    # ── Colour palette ──────────────────────────────────────────────────────
    DARK       = "1A1A2E"
    ACCENT     = "E94560"
    HEADER_FG  = "FFFFFF"
    ALT_ROW    = "F4F6FA"
    BORDER_CLR = "D0D5E0"

    thin   = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column definitions ──────────────────────────────────────────────────
    COLUMNS = [
        ("Log Type",            "log_type",          16),
        ("Exception Type",      "exception_type",    36),
        ("Frame Type",          "frame_type",        16),
        ("Frame Fingerprint",   "frame_fingerprint", 30),
        ("Current Thread",      "current_thread",    40),
        ("Top Frame Method",    "top_frame_method",  80),
        ("Group / Directory",   "group",             28),
    ]

    header_font  = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
    header_fill  = PatternFill("solid", fgColor=DARK)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = border
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 16

    body_font  = Font(name="Arial", size=11)
    alt_fill   = PatternFill("solid", fgColor=ALT_ROW)
    body_align = Alignment(vertical="top", wrap_text=True)
    err_font   = Font(name="Arial", size=11, color=ACCENT, bold=True)

    for row_idx, row in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            if isinstance(value, list):
                value = "\n".join(str(v) for v in value)
            elif value is None:
                value = ""

            cell           = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.border    = border
            cell.alignment = body_align
            cell.font      = err_font if key == "error" and value else body_font
            if fill.fill_type:
                cell.fill  = fill

        ws.row_dimensions[row_idx].height = 16

    # ── Freeze header & add auto-filter ────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Generated"
    ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A2"] = "Total groups"
    ws2["B2"] = len(rows)
    ws2["A3"] = "Errors"
    ws2["B3"] = sum(1 for r in rows if r.get("error"))
    for cell in [ws2["A1"], ws2["A2"], ws2["A3"]]:
        cell.font = Font(name="Arial", bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5050))
    print(f"🚀  Crash Analyser API  →  http://localhost:{port}")
    print(f"   Model : {DEFAULT_MODEL}")
    app.run(host="0.0.0.0", port=port, debug=False)
